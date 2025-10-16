#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试自定义标题功能
验证Excel下拉列表和ID列的标题可以从col_settings配置中自定义
"""

from openpyxl import load_workbook
import os
import sys
import importlib.util

# 添加当前目录到Python路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 直接导入temp.py中的函数
spec = importlib.util.spec_from_file_location("temp", os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "tests", "temp.py"))
temp = importlib.util.module_from_spec(spec)
spec.loader.exec_module(temp)


def test_custom_titles():
    """
    测试自定义标题功能
    创建一个Excel文件，使用自定义的下拉列表和ID列标题
    """
    try:
        print("\n==== 开始测试自定义标题功能 ====")
        
        # 定义测试配置，包含自定义标题
        mapping_config = {
            "mapping_data": {
                "水果映射表": {
                    "data": [
                        {'id': 101, 'name': '苹果'},
                        {'id': 102, 'name': '香蕉'},
                        {'id': 103, 'name': '橙子'},
                        {'id': 104, 'name': '西瓜'},
                        {'id': 105, 'name': '葡萄'}
                    ],
                    "id_col": 1,
                    "name_col": 2
                },
                "类型映射表": {
                    "data": [
                        {'id': 201, 'name': '温带水果'},
                        {'id': 202, 'name': '热带水果'},
                        {'id': 203, 'name': '柑橘类'},
                        {'id': 204, 'name': '浆果类'}
                    ],
                    "id_col": 1,
                    "name_col": 2
                }
            },
            "col_settings": [
                {
                    "dropdown_col": "A",
                    "id_col": "B",
                    "mapping_name": "水果映射表",
                    "title_row": 1,
                    "data_start_row": 2,
                    "data_end_row": 100,
                    "dropdown_title": "选择水果",  # 自定义下拉列表标题
                    "id_title": "水果ID"         # 自定义ID列标题
                },
                {
                    "dropdown_col": "C",
                    "id_col": "D",
                    "mapping_name": "类型映射表",
                    "title_row": 1,
                    "data_start_row": 2,
                    "data_end_row": 100,
                    "dropdown_title": "选择类型",  # 自定义下拉列表标题
                    "id_title": "类型ID"         # 自定义ID列标题
                }
            ]
        }
        
        # 创建包含自定义标题的Excel文件
        test_file_name = "custom_titles_excel.xlsx"
        temp.create_multiple_mappings_excel(mapping_config, test_file_name)
        
        # 验证生成的文件
        if not os.path.exists(test_file_name):
            raise FileNotFoundError(f"测试文件 '{test_file_name}' 未生成")
        
        print(f"\n已成功创建包含自定义标题的Excel文件: {test_file_name}")
        
        # 读取文件并验证标题设置
        wb = load_workbook(filename=test_file_name)
        ws_main = wb["数据录入"]
        
        # 验证自定义标题是否正确设置
        print("\n验证自定义标题设置:")
        
        # 检查第一组列的标题
        a1_title = ws_main.cell(row=1, column=1).value
        b1_title = ws_main.cell(row=1, column=2).value
        print(f"A1单元格标题: '{a1_title}' (期望: '选择水果')")
        print(f"B1单元格标题: '{b1_title}' (期望: '水果ID')")
        
        # 检查第二组列的标题
        c1_title = ws_main.cell(row=1, column=3).value
        d1_title = ws_main.cell(row=1, column=4).value
        print(f"C1单元格标题: '{c1_title}' (期望: '选择类型')")
        print(f"D1单元格标题: '{d1_title}' (期望: '类型ID')")
        
        # 验证标题是否符合预期
        titles_match = (a1_title == "选择水果" and 
                       b1_title == "水果ID" and 
                       c1_title == "选择类型" and 
                       d1_title == "类型ID")
        
        if titles_match:
            print("\n✅ 测试成功: 所有自定义标题都已正确设置")
        else:
            print("\n❌ 测试失败: 自定义标题设置不正确")
            raise AssertionError("自定义标题设置验证失败")
        
        # 显示成功消息
        print("\n==== 自定义标题功能测试完成 ====")
        print("功能总结:")
        print("1. 已成功在Excel文件中设置自定义的下拉列表和ID列标题")
        print("2. 支持为不同的列组设置不同的标题")
        print("3. 如果未提供自定义标题，将使用默认标题")
        print(f"\n请打开 '{test_file_name}' 文件查看自定义标题效果。")
        
        return True
    except Exception as e:
        print(f"\n❌ 测试失败: {str(e)}")
        raise


if __name__ == "__main__":
    test_custom_titles()