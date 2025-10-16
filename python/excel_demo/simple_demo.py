from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font


def simple_id_name_dropdown_demo(data_list, file_name="simple_dropdown_demo.xlsx"):
    """
    创建一个简化版的演示Excel文件，直观展示下拉列表选择名称并关联显示ID的功能
    
    参数:
        data_list (list): 包含id和name的字典列表，格式为[{'id': 123, 'name': 'xxx'}, ...]
        file_name (str): 保存的文件名
    
    返回值:
        str: 保存的文件名
    
    异常:
        Exception: 当创建或保存Excel文件失败时抛出
    """
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "演示表"
        
        # 添加标题行，使用加粗字体
        title_font = Font(bold=True)
        ws.cell(row=1, column=1).value = "选择名称 (下拉列表)"
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=2).value = "关联的ID (自动显示)"
        ws.cell(row=1, column=2).font = title_font
        
        # 创建隐藏的映射区域（在同一张表的右侧）
        # 这样更容易理解和调试
        mapping_start_col = 5
        mapping_start_row = 1
        
        # 设置映射区域标题
        ws.cell(row=mapping_start_row, column=mapping_start_col).value = "ID映射表（可隐藏）"
        ws.cell(row=mapping_start_row, column=mapping_start_col).font = title_font
        
        # 填充映射数据
        ws.cell(row=mapping_start_row+1, column=mapping_start_col).value = "ID"
        ws.cell(row=mapping_start_row+1, column=mapping_start_col+1).value = "名称"
        
        for idx, item in enumerate(data_list, start=2):
            ws.cell(row=mapping_start_row+idx, column=mapping_start_col).value = item['id']
            ws.cell(row=mapping_start_row+idx, column=mapping_start_col+1).value = item['name']
        
        # 获取映射数据范围
        mapping_name_range = f"$F${mapping_start_row+2}:$F${mapping_start_row+1+len(data_list)}"
        mapping_id_range = f"$E${mapping_start_row+2}:$E${mapping_start_row+1+len(data_list)}"
        
        # 设置下拉列表（在A列）
        dv = DataValidation(type="list", formula1=mapping_name_range)
        dv.add("A2:A100")  # 应用到A2到A100单元格
        ws.add_data_validation(dv)
        
        # 设置公式，根据选择的名称自动显示对应的ID（在B列）
        for row_idx in range(2, 101):
            # 使用INDEX和MATCH函数组合，从名称查找对应的ID
            formula = f'=IFERROR(INDEX({mapping_id_range}, MATCH(A{row_idx}, {mapping_name_range}, 0)), "")'
            ws.cell(row=row_idx, column=2).value = formula
        
        # 设置一些示例数据，让用户更直观地看到效果
        if data_list and len(data_list) >= 2:
            ws.cell(row=2, column=1).value = data_list[0]['name']  # 预先选择第一个数据
            ws.cell(row=3, column=1).value = data_list[1]['name']  # 预先选择第二个数据
        
        # 调整列宽，使内容更容易阅读
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        
        # 保存文件
        wb.save(file_name)
        
        print(f"\n简化演示版Excel文件 '{file_name}' 已成功创建！")
        print("功能特点：")
        print("1. A列（选择名称）设置了下拉列表，显示中文名称")
        print("2. B列（关联的ID）会自动根据选择的名称显示对应的ID值")
        print("3. 右侧E-F列为映射表，展示ID和名称的对应关系")
        print("4. 已预先在A2和A3单元格选择了示例数据，便于直观查看效果")
        print("\n使用方法：")
        print("- 点击A列的单元格（如A4），会显示下拉箭头")
        print("- 从下拉列表中选择一个名称")
        print("- B列会自动显示对应的ID值")
        print("- 如需隐藏映射表，可右键点击E-F列标题，选择'隐藏'")
        
        return file_name
    except Exception as e:
        print(f"创建演示Excel文件失败: {str(e)}")
        raise


if __name__ == "__main__":
    # 示例数据
    demo_data = [
        {"id": 101, "name": "苹果"},
        {"id": 102, "name": "香蕉"},
        {"id": 103, "name": "橙子"},
        {"id": 104, "name": "西瓜"},
        {"id": 105, "name": "葡萄"}
    ]
    
    # 运行简化版演示
    simple_id_name_dropdown_demo(demo_data)