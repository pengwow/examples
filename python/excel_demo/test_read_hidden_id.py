from openpyxl import load_workbook


def read_excel_with_hidden_data(file_name):
    """
    读取Excel文件，包括隐藏的数据
    
    参数:
        file_name (str): Excel文件名
    
    返回值:
        list: 包含读取的数据的列表
    
    异常:
        Exception: 当读取Excel文件失败时抛出
    """
    try:
        # 加载Excel文件
        wb = load_workbook(filename=file_name, data_only=True)  # data_only=True表示读取公式计算后的值
        
        # 获取主工作表
        ws_main = wb["数据录入"]
        
        # 尝试访问隐藏的映射表
        ws_mapping = wb["映射表"]
        
        # 读取主工作表中的数据
        print(f"\n正在读取Excel文件 '{file_name}' 中的数据...")
        print("注意：虽然B列在Excel中看起来是空白的，但程序可以读取到其中的ID值")
        
        results = []
        
        # 读取A列和B列的数据（跳过标题行）
        for row_idx in range(2, 101):
            name_cell = ws_main.cell(row=row_idx, column=1)
            id_cell = ws_main.cell(row=row_idx, column=2)
            
            # 只处理有数据的行
            if name_cell.value:
                data = {
                    'row': row_idx,
                    'name': name_cell.value,
                    'id': id_cell.value
                }
                results.append(data)
                print(f"行 {row_idx}: 名称='{name_cell.value}', ID='{id_cell.value}'")
        
        # 显示映射表中的完整数据关系
        print("\n映射表中的完整ID-名称对应关系:")
        for row_idx in range(2, ws_mapping.max_row + 1):
            id_value = ws_mapping.cell(row=row_idx, column=1).value
            name_value = ws_mapping.cell(row=row_idx, column=2).value
            print(f"ID: {id_value}, 名称: {name_value}")
        
        return results
    except Exception as e:
        print(f"读取Excel文件失败: {str(e)}")
        raise


if __name__ == "__main__":
    # 读取我们刚创建的Excel文件
    read_excel_with_hidden_data("dropdown_with_hidden_id.xlsx")