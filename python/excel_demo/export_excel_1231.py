# 根据A=[{"hostname":"192.168.1.1","ip":"192.168.1.1","mac":"00:11:22:33:44:55"}]\n# 导出excel表格文件,文件头为hostname,ip,mac,动态添加A中的所有key,这只是例子\n# 从第二行开始是A中的所有值

import openpyxl
from openpyxl.utils import get_column_letter
import csv
from io import BytesIO

def export_to_excel(data_list, output_file=None):
    """
    将字典列表导出到Excel文件或BytesIO对象
    :param data_list: 字典列表，如[{"hostname":"192.168.1.1","ip":"192.168.1.1","mac":"00:11:22:33:44:55"}]
    :param output_file: 输出Excel文件名或BytesIO对象。如果为None，则返回BytesIO对象
    :return: 如果output_file为None，则返回包含Excel数据的BytesIO对象
    """
    # 创建工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # 获取所有唯一的key作为表头
    if not data_list:
        print("数据列表为空")
        return
    
    # 动态获取所有key
    all_keys = set()
    for item in data_list:
        all_keys.update(item.keys())
    
    # 将key排序作为表头
    headers = sorted(list(all_keys))
    
    # 写入表头（第一行）
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        sheet[f"{col_letter}1"] = header
    
    # 写入数据（从第二行开始）
    for row_idx, item in enumerate(data_list, 2):
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            sheet[f"{col_letter}{row_idx}"] = item.get(header, "")
    
    # 保存到文件或BytesIO对象
    if output_file is None:
        # 返回BytesIO对象
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)  # 将指针移到开始位置
        print("Excel数据已导出到BytesIO对象")
        return buffer
    elif hasattr(output_file, 'write'):
        # 是文件对象或BytesIO对象
        workbook.save(output_file)
        output_file.seek(0)  # 将指针移到开始位置
        print("Excel数据已导出到文件对象")
        return output_file
    else:
        # 是文件路径
        workbook.save(output_file)
        print(f"Excel文件已导出：{output_file}")

def import_from_excel(input_file):
    """
    从Excel文件或文件对象导入数据并转换为字典列表
    :param input_file: 输入Excel文件名或文件对象（如BytesIO）
    :return: 字典列表，如[{"hostname":"192.168.1.1","ip":"192.168.1.1","mac":"00:11:22:33:44:55"}]
    """
    # 打开Excel文件（支持文件路径或文件对象）
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active
    
    # 获取所有数据行
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("Excel数据为空")
        return []
    
    # 获取表头（第一行）
    headers = list(rows[0])
    
    # 从第二行开始读取数据
    data_list = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue  # 跳过空行
        
        # 创建字典，将表头与对应的值关联
        item = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                if header:  # 跳过空表头
                    item[str(header)] = value if value is not None else ""
        
        data_list.append(item)
    
    print(f"已从Excel导入 {len(data_list)} 条数据")
    return data_list

def export_to_csv(data_list, output_file):
    """
    将字典列表导出到CSV文件（可被Excel打开）
    :param data_list: 字典列表，如[{"hostname":"192.168.1.1","ip":"192.168.1.1","mac":"00:11:22:33:44:55"}]
    :param output_file: 输出CSV文件名
    """
    # 获取所有唯一的key作为表头
    if not data_list:
        print("数据列表为空")
        return
    
    # 动态获取所有key（考虑到不同字典可能有不同的key）
    all_keys = set()
    for item in data_list:
        all_keys.update(item.keys())
    
    # 将key排序作为表头（可选）
    headers = sorted(list(all_keys))
    
    # 写入CSV文件
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        # 创建writer对象
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        
        # 写入表头（第一行）
        writer.writeheader()
        
        # 写入数据（从第二行开始）
        for item in data_list:
            writer.writerow(item)
    
    print(f"CSV文件已导出：{output_file}（可直接用Excel打开）")

# 示例使用
if __name__ == "__main__":
    # 示例数据
    A = [
        {"hostname":"192.168.1.1","ip":"192.168.1.1","mac":"00:11:22:33:44:55"},
        {"hostname":"192.168.1.2","ip":"192.168.1.2","mac":"00:11:22:33:44:56"},
        {"hostname":"192.168.1.3","ip":"192.168.1.3","mac":"00:11:22:33:44:57","status":"active"}  # 包含额外字段
    ]
    
    print("=== 1. 导出到文件测试 ===")
    # 导出到Excel文件
    export_to_excel(A, "output.xlsx")
    
    print("\n=== 2. 从文件导入测试 ===")
    # 导入Excel文件
    imported_data = import_from_excel("output.xlsx")
    
    print("\n=== 3. IO流功能测试 (导出) ===")
    # 导出到BytesIO对象（模拟返回给前端）
    excel_buffer = export_to_excel(A)
    print(f"BytesIO对象大小: {excel_buffer.getbuffer().nbytes} 字节")
    
    print("\n=== 4. IO流功能测试 (导入) ===")
    # 从BytesIO对象导入（模拟前端上传）
    io_imported_data = import_from_excel(excel_buffer)
    
    print("\n=== 5. IO流对象传递测试 ===")
    # 使用BytesIO对象作为参数传递
    import io
    file_obj = io.BytesIO()
    export_to_excel(A, file_obj)
    file_obj.seek(0)
    obj_imported_data = import_from_excel(file_obj)
    
    print("\n=== 6. 数据完整性验证 ===")
    print(f"原数据条数: {len(A)}")
    print(f"文件导入数据条数: {len(imported_data)}")
    print(f"IO导出→导入数据条数: {len(io_imported_data)}")
    print(f"文件对象传递数据条数: {len(obj_imported_data)}")
    
    # 检查所有导入方式是否得到相同结果
    all_same = len(imported_data) == len(io_imported_data) == len(obj_imported_data) == len(A)
    print(f"所有导入方式结果一致: {'✅' if all_same else '❌'}")
    
    # 也可以选择导出到CSV（保留作为备选）
    # export_to_csv(A, "output.csv")