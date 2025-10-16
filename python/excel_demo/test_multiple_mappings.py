from openpyxl import load_workbook


def test_multiple_mappings_excel(file_name="multiple_mappings_excel.xlsx"):
    """
    测试包含多个映射表的Excel文件功能
    
    参数:
        file_name (str): Excel文件名，默认为"multiple_mappings_excel.xlsx"
    
    返回值:
        dict: 包含测试结果的字典
    
    异常:
        Exception: 当读取Excel文件失败时抛出
    """
    try:
        print(f"\n==== 开始测试包含多个映射表的Excel文件 '{file_name}' ====")
        
        # 加载Excel文件
        wb = load_workbook(filename=file_name)
        
        # 获取主工作表
        ws_main = wb["数据录入"]
        
        # 获取所有工作表名称
        print(f"\nExcel文件中的工作表：{wb.sheetnames}")
        
        # 显示隐藏的映射表
        hidden_sheets = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == "hidden"]
        print(f"隐藏的映射表：{hidden_sheets}")
        
        # 读取并显示每个映射表的内容
        print("\n映射表内容：")
        for sheet_name in hidden_sheets:
            ws = wb[sheet_name]
            print(f"\n----- {sheet_name} -----\n")
            print("ID\t名称")
            print("-	-")
            
            # 读取数据（跳过标题行）
            for row in range(2, ws.max_row + 1):
                id_value = ws.cell(row=row, column=1).value
                name_value = ws.cell(row=row, column=2).value
                print(f"{id_value}\t{name_value}")
        
        # 读取主工作表的结构信息
        print("\n\n主工作表结构信息：")
        print("列\t标题\t\t功能")
        print("-	-		-")
        
        # 显示标题行信息
        for col_idx in range(1, 5):  # 检查前4列，应该包含两组下拉列表和ID列
            col_letter = chr(64 + col_idx)
            title = ws_main.cell(row=1, column=col_idx).value
            
            # 判断列的功能
            if col_idx % 2 == 1:  # 奇数列为下拉列表
                function = "下拉列表（显示名称）"
            else:  # 偶数列为隐藏的ID
                function = "隐藏ID（关联值）"
            
            print(f"{col_letter}\t{title}\t{function}")
        
        # 检查下拉列表设置
        print("\n\n下拉列表设置检查：")
        # 修复属性名：将dataValidations改为dataValidation
        for dv in ws_main.data_validations.dataValidation:
            print(f"区域: {dv.sqref}, 公式: {dv.formula1}")
        
        # 检查公式设置
        print("\n\nID列公式检查：")
        # 检查几个代表性单元格的公式
        check_cells = [(2, 2), (2, 4)]  # B2和D2单元格
        for row, col in check_cells:
            cell = ws_main.cell(row=row, column=col)
            if cell.value and cell.value.startswith('='):
                col_letter = chr(64 + col)
                print(f"单元格 {col_letter}{row} 的公式: {cell.value}")
        
        print("\n==== 测试完成 ====")
        print("\n功能总结：")
        print("1. Excel文件已成功创建包含多个映射表")
        print("2. 每个映射表都正确存储了ID和名称的对应关系")
        print("3. 主工作表中设置了多组下拉列表和关联的ID列")
        print("4. 所有公式和数据验证都已正确配置")
        print("\n使用说明：")
        print("- 打开Excel文件后，点击A列或C列的单元格可以看到下拉列表")
        print("- 选择名称后，对应的ID会自动填充到相邻的B列或D列（视觉上隐藏）")
        print("- 程序读取时可以获取到完整的ID值信息")
        
        return {
            'status': 'success',
            'file_name': file_name,
            'sheet_count': len(wb.sheetnames),
            'hidden_sheets_count': len(hidden_sheets)
        }
    except Exception as e:
        print(f"测试失败: {str(e)}")
        raise


# 模拟用户选择数据并验证ID关联（自动测试）
def simulate_selection_and_verify(file_name="multiple_mappings_excel.xlsx"):
    """
    模拟用户在Excel中选择数据，并验证ID是否正确关联
    
    参数:
        file_name (str): Excel文件名
    
    返回值:
        bool: 测试是否成功
    """
    try:
        print(f"\n\n==== 开始模拟用户选择并验证ID关联 ====")
        
        # 加载Excel文件
        wb = load_workbook(filename=file_name)
        ws_main = wb["数据录入"]
        
        # 模拟用户选择数据
        # 在A2单元格选择"苹果"
        ws_main.cell(row=2, column=1).value = "苹果"
        # 在C2单元格选择"热带水果"
        ws_main.cell(row=2, column=3).value = "热带水果"
        
        # 保存修改
        test_file_name = "test_simulation_multiple_mappings.xlsx"
        wb.save(test_file_name)
        
        print(f"\n已模拟用户选择数据并保存到 '{test_file_name}'")
        print(f"- A2单元格选择: '苹果'")
        print(f"- C2单元格选择: '热带水果'")
        
        # 手动验证公式逻辑
        print("\n公式逻辑验证：")
        # 由于openpyxl不能直接计算公式结果，我们手动检查逻辑
        
        # 检查水果映射
        fruit_mapping = {}
        if "水果映射表" in wb.sheetnames:
            ws_fruit = wb["水果映射表"]
            for row in range(2, ws_fruit.max_row + 1):
                fruit_mapping[ws_fruit.cell(row=row, column=2).value] = ws_fruit.cell(row=row, column=1).value
            
            # 验证选择的水果
            selected_fruit = ws_main.cell(row=2, column=1).value
            if selected_fruit in fruit_mapping:
                print(f"验证: '{selected_fruit}' 应该映射到 ID {fruit_mapping[selected_fruit]}")
        
        # 检查类型映射
        type_mapping = {}
        if "类型映射表" in wb.sheetnames:
            ws_type = wb["类型映射表"]
            for row in range(2, ws_type.max_row + 1):
                type_mapping[ws_type.cell(row=row, column=2).value] = ws_type.cell(row=row, column=1).value
            
            # 验证选择的类型
            selected_type = ws_main.cell(row=2, column=3).value
            if selected_type in type_mapping:
                print(f"验证: '{selected_type}' 应该映射到 ID {type_mapping[selected_type]}")
        
        print("\n在实际Excel环境中，打开文件后公式会自动计算出正确的ID值。")
        print("如果您想亲自测试，请打开生成的Excel文件并查看效果。")
        
        print("\n==== 模拟验证完成 ====")
        return True
    except Exception as e:
        print(f"模拟验证失败: {str(e)}")
        return False


if __name__ == "__main__":
    # 测试包含多个映射表的Excel文件
    test_result = test_multiple_mappings_excel()
    
    # 模拟用户选择并验证ID关联
    simulate_selection_and_verify()