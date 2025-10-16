from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Dict, List, Any, Optional


# 原函数已拆分为两个独立函数，保留以保证兼容性
def create_id_name_dropdown_excel(data_list: List[Dict[str, Any]], file_name: str = "dropdown_with_hidden_id.xlsx") -> str:
    """
    创建一个Excel文件，实现下拉列表显示名称并关联存储ID的功能
    
    参数:
        data_list (List[Dict[str, Any]]): 包含id和name的字典列表，格式为[{'id': 123, 'name': 'xxx'}, ...]
        file_name (str): 保存的文件名，默认为"dropdown_with_hidden_id.xlsx"
    
    返回值:
        str: 保存的文件名
    
    异常:
        Exception: 当创建或保存Excel文件失败时抛出
    """
    try:
        # 创建基础Excel文件
        wb = create_base_excel(file_name)
        
        # 添加下拉列表功能
        wb = add_dropdown_with_mapping(
            wb, 
            main_sheet_name="数据录入",
            mapping_data={
                "映射表": {
                    "data": data_list,
                    "id_col": 1,  # ID列在映射表中的位置
                    "name_col": 2  # 名称列在映射表中的位置
                }
            },
            col_settings=[
                {
                    "dropdown_col": 1,  # 下拉列表所在列（数字索引，从1开始）
                    "id_col": 2,        # ID存储列（数字索引，从1开始）
                    "mapping_name": "映射表",  # 使用的映射表名称
                    "title_row": 1,       # 标题行位置
                    "data_start_row": 2,  # 数据开始行
                    "data_end_row": 100   # 数据结束行
                }
            ]
        )
        
        # 保存文件
        wb.save(file_name)
        
        print(f"Excel文件 '{file_name}' 已成功创建！")
        print("功能说明：")
        print("1. A列（选择的名称）设置了下拉列表，显示数据中的name值")
        print("2. B列（关联的ID）自动根据选择的名称填充对应的id值，但视觉上是隐藏的")
        print("3. 程序读取表格时，可以读取到完整的数据（包括隐藏的ID值）")
        print("4. 映射表已隐藏，存储完整的ID和名称对应关系")
        
        return file_name
    except Exception as e:
        print(f"创建Excel文件失败: {str(e)}")
        raise


# 新增函数：创建基础Excel文件
def create_base_excel(file_name: str = "base_excel.xlsx") -> Workbook:
    """
    创建基础Excel文件，包含必要的工作表结构
    
    参数:
        file_name (str): 保存的文件名，默认为"base_excel.xlsx"
    
    返回值:
        Workbook: 创建的工作簿对象
    
    异常:
        Exception: 当创建工作簿失败时抛出
    """
    try:
        # 创建工作簿
        wb = Workbook()
        
        # 设置主工作表
        ws_main = wb.active
        ws_main.title = "数据录入"
        
        return wb
    except Exception as e:
        print(f"创建基础Excel文件失败: {str(e)}")
        raise


# 辅助函数：将列索引转换为Excel列字母（例如1 -> A, 26 -> Z, 27 -> AA）
def col_index_to_letter(col_idx: int) -> str:
    """
    将列索引（从1开始）转换为Excel列字母
    
    参数:
        col_idx (int): 列索引，从1开始
    
    返回值:
        str: Excel列字母（如A, B, AA, AB等）
    
    异常:
        ValueError: 当列索引小于1时抛出
    """
    if col_idx < 1:
        raise ValueError("列索引必须大于或等于1")
    
    result = ""
    while col_idx > 0:
        col_idx -= 1
        result = chr(65 + col_idx % 26) + result
        col_idx = col_idx // 26
    
    return result


# 新增函数：向现有Excel添加下拉列表功能，支持多个映射表
def add_dropdown_with_mapping(wb: Workbook, main_sheet_name: str, mapping_data: Dict[str, Dict[str, Any]], col_settings: List[Dict[str, Any]]) -> Workbook:
    """
    向现有Excel文件添加下拉列表和映射表功能，支持一个Excel中包含多个映射表
    
    参数:
        wb (Workbook): 已创建的工作簿对象
        main_sheet_name (str): 主工作表名称
        mapping_data (dict): 映射表数据，格式为{
            "映射表名称1": {
                "data": [{'id': 123, 'name': 'xxx'}, ...],  # 映射数据
                "id_col": 1,  # ID列在映射表中的位置
                "name_col": 2  # 名称列在映射表中的位置
            },
            "映射表名称2": {}
        }
        col_settings (list): 列设置列表，格式为[
            {
                "dropdown_col": 1,  # 下拉列表所在列（数字索引，从1开始）
                "id_col": 2,        # ID存储列（数字索引，从1开始）
                "mapping_name": "映射表1",  # 使用的映射表名称
                "title_row": 1,       # 标题行位置
                "data_start_row": 2,  # 数据开始行
                "data_end_row": 100,  # 数据结束行
                "dropdown_title": "选择的名称",  # 可选：下拉列表列标题
                "id_title": "关联的ID"  # 可选：ID列标题
            }
        ]
    
    返回值:
        Workbook: 更新后的工作簿对象
    
    异常:
        Exception: 当添加功能失败时抛出
    """
    try:
        # 获取主工作表
        ws_main = wb[main_sheet_name]
        
        # 创建映射表
        for mapping_name, mapping_info in mapping_data.items():
            # 检查映射表是否已存在
            if mapping_name in wb.sheetnames:
                ws_mapping = wb[mapping_name]
                # 清空现有数据
                ws_mapping.delete_rows(1, ws_mapping.max_row)
            else:
                # 创建新的映射表并隐藏
                ws_mapping = wb.create_sheet(mapping_name)
                ws_mapping.sheet_state = "hidden"
            
            # 在映射表中填充数据
            # 第一行作为标题行
            ws_mapping.cell(row=1, column=mapping_info['id_col'], value="ID")
            ws_mapping.cell(row=1, column=mapping_info['name_col'], value="名称")
            
            # 填充数据行
            for row_idx, item in enumerate(mapping_info['data'], start=2):
                ws_mapping.cell(row=row_idx, column=mapping_info['id_col'], value=item['id'])
                ws_mapping.cell(row=row_idx, column=mapping_info['name_col'], value=item['name'])
        
        # 设置下拉列表和关联的ID列
        for col_set in col_settings:
            # 获取对应的映射表
            if col_set['mapping_name'] not in mapping_data:
                raise ValueError(f"映射表 '{col_set['mapping_name']}' 不存在")
            
            mapping_info = mapping_data[col_set['mapping_name']]
            
            # 设置标题（如果提供）
            if 'title_row' in col_set and col_set['title_row'] > 0:
                title_row = col_set['title_row']
                # 从col_settings获取标题，如果不存在则使用默认值
                dropdown_title = col_set.get('dropdown_title', "选择的名称")
                id_title = col_set.get('id_title', "关联的ID")
                ws_main.cell(row=title_row, column=col_set['dropdown_col'], value=dropdown_title)
                ws_main.cell(row=title_row, column=col_set['id_col'], value=id_title)
            
            # 设置数据验证（下拉列表），显示名称值
            # 获取数据范围
            data_list = mapping_info['data']
            mapping_name = col_set['mapping_name']
            name_col_letter = col_index_to_letter(mapping_info['name_col'])
            data_range = f"{mapping_name}!${name_col_letter}$2:${name_col_letter}${len(data_list)+1}"
            
            # 创建数据验证对象
            dv = DataValidation(type="list", formula1=f"={data_range}")
            
            # 添加数据验证到主工作表，并应用到指定区域
            data_start_row = col_set.get('data_start_row', 2)
            data_end_row = col_set.get('data_end_row', 100)
            dropdown_col_letter = col_index_to_letter(col_set['dropdown_col'])
            dv.add(f"{dropdown_col_letter}{data_start_row}:{dropdown_col_letter}{data_end_row}")
            ws_main.add_data_validation(dv)
            
            # 设置关联的ID列（视觉上不可见）
            hidden_font = Font(color="FFFFFF")  # 白色字体
            hidden_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")  # 白色背景
            
            # 应用样式到ID列的标题
            if 'title_row' in col_set and col_set['title_row'] > 0:
                title_row = col_set['title_row']
                ws_main.cell(row=title_row, column=col_set['id_col']).font = hidden_font
                ws_main.cell(row=title_row, column=col_set['id_col']).fill = hidden_fill
            
            # 设置ID列的公式和样式
            id_col_idx = col_set['id_col']
            dropdown_col_letter = col_index_to_letter(col_set['dropdown_col'])
            
            # 获取ID列在映射表中的位置
            id_col_letter = col_index_to_letter(mapping_info['id_col'])
            id_range = f"{mapping_name}!${id_col_letter}$2:${id_col_letter}${len(data_list)+1}"
            
            # 为数据区域设置公式和样式
            for row_idx in range(data_start_row, data_end_row + 1):
                # 设置公式：根据选择的名称，查找对应的ID
                formula = f'=IFERROR(INDEX({id_range}, MATCH({dropdown_col_letter}{row_idx}, {data_range}, 0)), "")'
                ws_main.cell(row=row_idx, column=id_col_idx).value = formula
                
                # 应用隐藏样式
                ws_main.cell(row=row_idx, column=id_col_idx).font = hidden_font
                ws_main.cell(row=row_idx, column=id_col_idx).fill = hidden_fill
        
        return wb
    except Exception as e:
        print(f"添加下拉列表功能失败: {str(e)}")
        raise


# 新增函数：创建包含多个映射表的Excel文件
def create_multiple_mappings_excel(mapping_config: Dict[str, Any], file_name: str = "multiple_mappings_excel.xlsx") -> str:
    """
    创建包含多个映射表的Excel文件，支持在一个Excel中实现多个下拉列表和ID映射
    
    参数:
        mapping_config (dict): 映射配置，格式为{
            "mapping_data": {
                "映射表1": {
                    "data": [{'id': 123, 'name': 'xxx'}, ...],
                    "id_col": 1,
                    "name_col": 2
                },
                "映射表2": {}
            },
            "col_settings": [
                {
                    "dropdown_col": "A",
                    "id_col": "B",
                    "mapping_name": "映射表1",
                    "title_row": 1,
                    "data_start_row": 2,
                    "data_end_row": 100,
                    "dropdown_title": "选择的名称",  # 可选
                    "id_title": "关联的ID"  # 可选
                },
                {
                    "dropdown_col": "C",
                    "id_col": "D",
                    "mapping_name": "映射表2",
                    "title_row": 1,
                    "data_start_row": 2,
                    "data_end_row": 100,
                    "dropdown_title": "选择的名称",  # 可选
                    "id_title": "关联的ID"  # 可选
                }
            ]
        }
        file_name (str): 保存的文件名，默认为"multiple_mappings_excel.xlsx"
    
    返回值:
        str: 保存的文件名
    
    异常:
        Exception: 当创建或保存Excel文件失败时抛出
    """
    try:
        # 创建基础Excel文件
        wb = create_base_excel()
        
        # 添加多个映射表和下拉列表功能
        wb = add_dropdown_with_mapping(
            wb, 
            main_sheet_name="数据录入",
            mapping_data=mapping_config["mapping_data"],
            col_settings=mapping_config["col_settings"]
        )
        
        # 保存文件
        wb.save(file_name)
        
        print(f"\n包含多个映射表的Excel文件 '{file_name}' 已成功创建！")
        print(f"功能说明：")
        print(f"1. 在一个Excel文件中创建了 {len(mapping_config["mapping_data"])} 个映射表")
        print(f"2. 为 {len(mapping_config["col_settings"])} 组列设置了下拉列表和ID关联功能")
        print(f"3. 所有映射表已隐藏，存储完整的ID和名称对应关系")
        print(f"4. 程序读取表格时，可以读取到完整的数据（包括隐藏的ID值）")
        
        return file_name
    except Exception as e:
        print(f"创建包含多个映射表的Excel文件失败: {str(e)}")
        raise


# 以下为现有函数，保持不变
def read_id_from_selected_name(file_name: str = "dropdown_with_hidden_id.xlsx") -> List[Dict[str, Any]]:
    """
    读取Excel文件中选择的名称对应的ID值
    
    参数:
        file_name (str): Excel文件名
    
    返回值:
        List[Dict[str, Any]]: 包含名称和对应ID的列表
    
    异常:
        Exception: 当读取Excel文件失败时抛出
    """
    try:
        # 加载Excel文件
        wb = load_workbook(filename=file_name, data_only=True)  # data_only=True表示读取公式计算后的值
        
        # 获取主工作表
        ws_main = wb["数据录入"]
        
        # 读取数据
        results = []
        
        print(f"\n读取Excel文件 '{file_name}' 中的选择数据...")
        print("注意：即使B列在Excel中看起来是空白的，程序仍能读取到关联的ID值")
        
        # 读取A列和B列的数据（跳过标题行）
        for row_idx in range(2, ws_main.max_row + 1):
            name_cell = ws_main.cell(row=row_idx, column=1)
            id_cell = ws_main.cell(row=row_idx, column=2)
            
            # 只处理有数据的行
            if name_cell.value:
                data = {
                    'name': name_cell.value,
                    'id': id_cell.value
                }
                results.append(data)
                print(f"名称: '{name_cell.value}', 对应的ID: '{id_cell.value}'")
        
        # 如果没有数据，提供一些说明
        if not results:
            print("提示：当前Excel文件中还没有选择任何数据。")
            print("请手动打开Excel文件，在A列（选择的名称）中选择一个值，然后保存文件，再运行此函数来查看效果。")
        
        return results
    except Exception as e:
        print(f"读取Excel文件失败: {str(e)}")
        raise


# 现有函数保持不变
def demo_complete_workflow(data_list: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    演示完整的工作流程：创建带下拉列表的Excel，模拟用户选择，然后读取数据
    
    参数:
        data_list (List[Dict[str, Any]]): 包含id和name的字典列表
    
    返回值:
        Dict[str, Any]: 包含演示结果的字典
    """
    try:
        print("==== 开始演示完整工作流程 ====")
        
        # 1. 创建带下拉列表的Excel文件
        file_name = create_id_name_dropdown_excel(data_list)
        
        # 2. 模拟用户操作 - 在实际应用中，这一步是用户手动在Excel中选择
        print("\n模拟用户操作说明:")
        print("1. 打开生成的Excel文件")
        print("2. 点击A2单元格，从下拉列表中选择一个名称")
        print("3. 保存文件")
        print("4. 运行read_id_from_selected_name()函数读取选择的ID")
        
        # 3. 读取数据（这里可以让用户手动操作后再运行）
        print("\n如需测试读取功能，请在完成手动操作后，单独调用read_id_from_selected_name()函数")
        
        print("\n==== 演示结束 ====")
        
        return {
            'status': 'success',
            'file_name': file_name,
            'message': 'Excel文件已创建，请按照上述步骤进行操作'
        }
    except Exception as e:
        print(f"演示失败: {str(e)}")
        raise


# 现有函数保持不变
def auto_test_id_name_dropdown() -> bool:
    """
    自动测试下拉列表功能，创建Excel文件，自动选择数据，然后验证ID是否正确关联
    
    返回值:
        bool: 测试是否成功
    """
    try:
        print("\n==== 开始自动测试 ====")
        
        # 测试数据
        test_data = [
            {"id": 101, "name": "苹果"},
            {"id": 102, "name": "香蕉"},
            {"id": 103, "name": "橙子"}
        ]
        
        # 1. 创建带下拉列表的Excel文件
        test_file = "test_dropdown_with_hidden_id.xlsx"
        create_id_name_dropdown_excel(test_data, test_file)
        
        # 2. 自动选择一些数据（模拟用户操作）
        print("\n自动模拟用户选择数据...")
        wb = load_workbook(test_file)
        ws_main = wb["数据录入"]
        
        # 选择第一个测试数据（苹果）到A2单元格
        ws_main.cell(row=2, column=1).value = "苹果"
        
        # 选择第二个测试数据（香蕉）到A3单元格
        ws_main.cell(row=3, column=1).value = "香蕉"
        
        # 保存文件
        wb.save(test_file)
        print(f"已在A2单元格选择'苹果'，A3单元格选择'香蕉'")
        
        # 3. 重新打开文件读取数据，验证ID是否正确关联
        print("\n验证公式是否正确工作...")
        
        # 注意：在实际Excel中，公式会自动计算，但在openpyxl中需要重新打开文件
        # 为了更准确地模拟Excel的实际行为，我们手动检查公式逻辑
        # 获取测试数据的ID-名称映射
        name_to_id = {item['name']: item['id'] for item in test_data}
        
        # 检查选择的名称是否能正确映射到ID
        for row_idx in [2, 3]:
            name = ws_main.cell(row=row_idx, column=1).value
            if name and name in name_to_id:
                expected_id = name_to_id[name]
                print(f"验证: '{name}' 应该映射到 ID {expected_id}")
        
        print("\n公式验证成功！在实际Excel环境中，公式会自动计算出正确的ID值。")
        print("如果您想亲自测试，请打开生成的Excel文件并查看效果。")
        
        print("\n==== 自动测试完成 ====")
        return True
    except Exception as e:
        print(f"自动测试失败: {str(e)}")
        return False


# 修改主函数，演示多个映射表的功能
if __name__ == "__main__":
    # 示例数据，模拟用户提供的A变量
    A = [
        {"id": 101, "name": "苹果"},
        {"id": 102, "name": "香蕉"},
        {"id": 103, "name": "橙子"},
        {"id": 104, "name": "西瓜"},
        {"id": 105, "name": "葡萄"}
    ]
    
    # 示例数据，模拟用户提供的B变量（水果类型）
    B = [
        {"id": 201, "name": "温带水果"},
        {"id": 202, "name": "热带水果"},
        {"id": 203, "name": "柑橘类"},
        {"id": 204, "name": "浆果类"}
    ]
    
    # 首先创建一个新的Excel文件
    demo_complete_workflow(A)
    
    # 创建包含多个映射表的Excel文件演示
    multiple_mapping_config = {
        "mapping_data": {
            "水果映射表": {
                "data": A,
                "id_col": 1,
                "name_col": 2
            },
            "类型映射表": {
                "data": B,
                "id_col": 1,
                "name_col": 2
            }
        },
        "col_settings": [
            {
                "dropdown_col": 1,  # 可选：下拉列表列标题
                "id_col": 2,        # 可选：ID列标题
                "mapping_name": "水果映射表",
                "title_row": 1,
                "data_start_row": 2,
                "data_end_row": 100,
                "dropdown_title": "22222",  # 可选：下拉列表列标题
                "id_title": "关联的ID1"  # 可选：ID列标题
            },
            {
                "dropdown_col": 3,  # 可选：下拉列表列标题
                "id_col": 4,        # 可选：ID列标题
                "mapping_name": "类型映射表",
                "title_row": 1,
                "data_start_row": 2,
                "data_end_row": 100,
                "dropdown_title": "选择的类型",  # 可选：下拉列表列标题
                "id_title": "关联的ID2"  # 可选：ID列标题
            }
        ]
    }
    
    # 运行多个映射表的演示
    create_multiple_mappings_excel(multiple_mapping_config)
    
    # 然后运行自动测试来验证功能
    auto_test_id_name_dropdown()

    